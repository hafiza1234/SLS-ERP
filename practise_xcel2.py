from openpyxl import load_workbook
wb = load_workbook("Book1.xlsx")
sheet = wb.worksheets[0]
a = sheet.cell(row=1,column=1)
sheet.cell(row=1,column=1).value = "name"
b = sheet.cell(row=2,column=2)
sheet.cell(row=2,column=2).value = "Address"
c = sheet.cell(row=3,column=3)
sheet.cell(row=3,column=3).value = "designation"
wb.save("Book1.xlsx")
